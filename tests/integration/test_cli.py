import csv as csv_module
import json
from pathlib import Path
import zipfile

import duckdb
from openpyxl import Workbook
from openpyxl import load_workbook
import pytest
from typer.testing import CliRunner

from gt_dataslicer.cli import app, create_app


runner = CliRunner()


def write_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "CD_EMPRESA,ST_CONTRATO,CD_NATUREZA_OPERACAO,CD_MODALIDADE,VALOR_TOTAL,DATA_ADMISSAO,NOME,CPF,STATUS,CIDADE,TIPO",
                "1,A,13,10,1500,2021-05-01,JOAO SILVA,123,ATIVO,SAO PAULO,FIXO",
                "1,P,13,10,2000,2021-05-02,MARIA SILVA,456,ATIVO,SAO PAULO,FIXO",
                "2,A,14,8,900,2019-01-01,ANA LIMA,,SUSPENSO,RIO,TEMPORARIO",
                "1,A,16,14,1200,2022-01-01,PEDRO,789,CANCELADO,SAO CARLOS,FIXO",
            ]
        ),
        encoding="utf-8",
    )


def rows_from_xlsx(path: Path, sheet: str = "Results_001") -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True)
    worksheet = workbook[sheet]
    return [tuple(row) for row in worksheet.iter_rows(values_only=True)]


def rows_from_csv(path: Path) -> list[tuple[str, ...]]:
    with path.open(newline="", encoding="utf-8") as file:
        return [tuple(row) for row in csv_module.reader(file)]


def rows_from_parquet(path: Path) -> list[tuple[object, ...]]:
    return duckdb.connect().execute(f"SELECT * FROM read_parquet('{path.as_posix()}')").fetchall()


def test_filter_command_exports_matching_rows(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    report_path = tmp_path / "report.json"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            'CD_EMPRESA = 1 AND ST_CONTRATO != "P" AND CD_NATUREZA_OPERACAO NOT IN (14, 15) AND CD_MODALIDADE <= 13',
            "--select",
            "NOME",
            "--select",
            "VALOR_TOTAL",
            "--report",
            str(report_path),
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_csv(output_path)
    assert rows == [("NOME", "VALOR_TOTAL"), ("JOAO SILVA", "1500")]
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["output_rows"] == 1
    assert report["output_paths"] == [str(output_path)]
    assert report["engine_options"]["output_format"] == "csv"
    assert report_path.exists()


def test_filter_command_accepts_pt_br_expression(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            'CD_EMPRESA = 1 E ST_CONTRATO != "P" E CD_NATUREZA_OPERACAO NAO EM (14, 15) E CD_MODALIDADE <= 13',
            "--select",
            "NOME",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("NOME",), ("JOAO SILVA",)]


def test_filter_command_accepts_pt_br_command_and_options(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS EM ("ATIVO", "SUSPENSO")',
            "--selecionar",
            "NOME",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("NOME",), ("JOAO SILVA",), ("MARIA SILVA",), ("ANA LIMA",)]


def test_filter_command_supports_parquet_input(tmp_path: Path) -> None:
    parquet_path = tmp_path / "input.parquet"
    output_path = tmp_path / "output.csv"
    duckdb.connect().execute(
        f"COPY (SELECT 1 AS ID, 'ATIVO' AS STATUS UNION ALL SELECT 2, 'CANCELADO') TO '{parquet_path.as_posix()}'"
        " (FORMAT parquet)"
    )

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(parquet_path),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS = "ATIVO"',
            "--selecionar",
            "ID",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("ID",), ("1",)]


def test_filter_command_supports_xlsx_input(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dados"
    sheet.append(["ID", "STATUS"])
    sheet.append([1, "ATIVO"])
    sheet.append([2, "CANCELADO"])
    xlsx_path = tmp_path / "input.xlsx"
    output_path = tmp_path / "output.csv"
    workbook.save(xlsx_path)

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(xlsx_path),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS = "ATIVO"',
            "--selecionar",
            "ID",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("ID",), ("1",)]


def test_filter_command_generates_summarization_and_summary_output(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    csv_path.write_text("STATUS,VALOR_TOTAL\nATIVO,10\nATIVO,30\nSUSPENSO,20\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS = "ATIVO"',
            "--sumarizacao",
            "--grupo-sumarizacao",
            "STATUS",
            "--totais-sumarizacao",
            "VALOR_TOTAL",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("STATUS", "VALOR_TOTAL"), ("ATIVO", "10"), ("ATIVO", "30")]
    assert rows_from_csv(tmp_path / "output_summary.csv") == [
        ("STATUS", "total_VALOR_TOTAL", "count"),
        ("ATIVO", "40.0", "2"),
    ]


def test_filter_command_summarization_only_generates_only_requested_output_file(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    csv_path.write_text("STATUS,VALOR_TOTAL\nATIVO,10\nATIVO,30\nSUSPENSO,20\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS IN ("ATIVO", "SUSPENSO")',
            "--somente-sumarizacao",
            "--grupo-sumarizacao",
            "STATUS",
            "--totais-sumarizacao",
            "VALOR_TOTAL",
        ],
    )

    assert result.exit_code == 0, result.output
    assert output_path.exists()
    summary_rows = rows_from_csv(output_path)
    assert summary_rows[0] == ("STATUS", "total_VALOR_TOTAL", "count")
    assert set(summary_rows[1:]) == {("ATIVO", "40.0", "2"), ("SUSPENSO", "20.0", "1")}
    assert not (tmp_path / "output_summary.csv").exists()
    assert not (tmp_path / "output.xlsx").exists()


def test_filter_command_processes_multiple_inputs_sequentially(tmp_path: Path) -> None:
    first = tmp_path / "jan.csv"
    second = tmp_path / "feb.csv"
    output_path = tmp_path / "filtered.csv"
    first.write_text("ID,STATUS\n1,ATIVO\n2,CANCELADO\n", encoding="utf-8")
    second.write_text("ID,STATUS\n3,ATIVO\n4,CANCELADO\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(first),
            str(second),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS = "ATIVO"',
            "--selecionar",
            "ID",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(tmp_path / "filtered_001_jan.csv") == [("ID",), ("1",)]
    assert rows_from_csv(tmp_path / "filtered_002_feb.csv") == [("ID",), ("3",)]
    assert not output_path.exists()


def test_filter_command_output_names_allow_new_directory_for_excel_sheets(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "input.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "Jan"
    first.append(["ID", "STATUS"])
    first.append([1, "ATIVO"])
    second = workbook.create_sheet("Feb")
    second.append(["ID", "STATUS"])
    second.append([2, "ATIVO"])
    workbook.save(xlsx_path)
    output_dir = tmp_path / "named_outputs"

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(xlsx_path),
            "--todas-abas",
            "--saida",
            str(output_dir),
            "--nome-saida",
            "janeiro",
            "--nome-saida",
            "fevereiro",
            "--filtro",
            'STATUS = "ATIVO"',
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_dir / "janeiro.csv") == [("ID", "STATUS"), ("1", "ATIVO")]
    assert rows_from_csv(output_dir / "fevereiro.csv") == [("ID", "STATUS"), ("2", "ATIVO")]


def test_filter_dry_run_fails_when_any_queue_input_fails(tmp_path: Path) -> None:
    first = tmp_path / "ok.csv"
    second = tmp_path / "bad.csv"
    output_path = tmp_path / "filtered.csv"
    first.write_text("ID,STATUS\n1,ATIVO\n", encoding="utf-8")
    second.write_text("ID,OTHER\n2,ATIVO\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(first),
            str(second),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS = "ATIVO"',
            "--teste",
        ],
    )

    assert result.exit_code == 1, result.output
    assert "não puderam ser processados" in result.output
    assert "Execução de teste concluída" not in result.output


def test_filter_report_includes_input_resolution_warnings(tmp_path: Path) -> None:
    zip_path = tmp_path / "inputs.zip"
    output_path = tmp_path / "filtered.csv"
    report_path = tmp_path / "report.json"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.writestr("data.csv", "A\n1\n")
        archive.writestr("ignored.txt", "not supported")

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(zip_path),
            "--saida",
            str(output_path),
            "--relatorio",
            str(report_path),
        ],
    )

    assert result.exit_code == 0, result.output
    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert report["input_path"] == f"{zip_path}!data.csv"
    assert report["warnings"] == ["Unsupported ZIP entry skipped: ignored.txt"]


def test_validate_filter_does_not_delete_zip_when_delete_flag_is_passed(tmp_path: Path) -> None:
    zip_path = tmp_path / "inputs.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.writestr("data.csv", "A\n1\n")

    result = runner.invoke(
        app,
        [
            "validar-filtro",
            str(zip_path),
            "--filtro",
            'A = "1"',
            "--excluir-zip-apos-extrair",
        ],
    )

    assert result.exit_code == 0, result.output
    assert zip_path.exists()


def test_filter_defaults_to_csv_when_output_has_no_suffix(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            'STATUS = "SUSPENSO"',
            "--select",
            "NOME",
        ],
    )

    normalized_path = output_path.with_suffix(".csv")
    assert result.exit_code == 0, result.output
    assert not output_path.exists()
    assert rows_from_csv(normalized_path) == [("NOME",), ("ANA LIMA",)]


def test_filter_writes_xlsx_when_suffix_requests_excel(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            'STATUS = "SUSPENSO"',
            "--select",
            "NOME",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_xlsx(output_path) == [("NOME",), ("ANA LIMA",)]


def test_filter_format_xlsx_adds_missing_suffix(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--format",
            "xlsx",
            "--where",
            'STATUS = "SUSPENSO"',
            "--select",
            "NOME",
        ],
    )

    normalized_path = output_path.with_suffix(".xlsx")
    assert result.exit_code == 0, result.output
    assert not output_path.exists()
    assert rows_from_xlsx(normalized_path) == [("NOME",), ("ANA LIMA",)]


def test_filter_writes_parquet_when_suffix_requests_parquet(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.parquet"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'STATUS = "SUSPENSO"',
            "--selecionar",
            "NOME",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_parquet(output_path) == [("ANA LIMA",)]


def test_filter_format_parquet_adds_missing_suffix(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--format",
            "parquet",
            "--where",
            'STATUS = "SUSPENSO"',
            "--select",
            "NOME",
        ],
    )

    normalized_path = output_path.with_suffix(".parquet")
    assert result.exit_code == 0, result.output
    assert not output_path.exists()
    assert rows_from_parquet(normalized_path) == [("ANA LIMA",)]


def test_filter_rejects_explicit_format_suffix_conflict(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--format",
            "csv",
            "--where",
            "CD_EMPRESA = 1",
        ],
    )

    assert result.exit_code == 2
    assert "conflita com o sufixo" in result.output
    assert not output_path.exists()


def test_filter_command_accepts_derived_column_json(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    write_csv(csv_path)
    derived = {
        "source": "CPF",
        "name": {"prefix": "LIMPO", "separator": "_"},
        "transforms": [{"operation": "keep_digits"}, {"operation": "add_prefix", "text": "BR_"}],
    }

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'NOME contains "JOAO"',
            "--selecionar",
            "NOME",
            "--coluna-derivada",
            json.dumps(derived),
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("NOME", "LIMPO_CPF"), ("JOAO SILVA", "BR_123")]


def test_filter_command_accepts_derived_columns_file(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    derived_path = tmp_path / "derived.yaml"
    write_csv(csv_path)
    derived_path.write_text(
        "derived_columns:\n"
        "  - source: CPF\n"
        "    name: CPF_FORMATADO\n"
        "    position:\n"
        "      mode: before\n"
        "      target: NOME\n"
        "    transforms:\n"
        "      - operation: format_cpf\n",
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'NOME contains "JOAO"',
            "--selecionar",
            "NOME",
            "--selecionar",
            "CPF",
            "--colunas-derivadas-arquivo",
            str(derived_path),
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("CPF_FORMATADO", "NOME", "CPF"), ("123", "JOAO SILVA", "123")]


def test_filter_command_accepts_derived_columns_from_config(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    config_path = tmp_path / "filters.json"
    write_csv(csv_path)
    config_path.write_text(
        json.dumps(
            {
                "where": 'NOME contains "JOAO"',
                "select": ["NOME"],
                "derived_columns": [
                    {
                        "source": "NOME",
                        "name": {"suffix": "MAIUSCULO", "separator": "_"},
                        "transforms": [{"operation": "uppercase"}],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--configuracao",
            str(config_path),
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("NOME", "NOME_MAIUSCULO"), ("JOAO SILVA", "JOAO SILVA")]


def test_xlsx_derived_columns_use_formula_with_cached_value_when_simple(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)
    derived = {
        "source": "NOME",
        "name": {"suffix": "MAIUSCULO", "separator": "_"},
        "transforms": [{"operation": "uppercase"}],
    }

    result = runner.invoke(
        app,
        [
            "filtrar",
            str(csv_path),
            "--saida",
            str(output_path),
            "--filtro",
            'NOME contains "JOAO"',
            "--selecionar",
            "NOME",
            "--formato",
            "xlsx",
            "--coluna-derivada",
            json.dumps(derived),
        ],
    )

    assert result.exit_code == 0, result.output
    formula_workbook = load_workbook(output_path, data_only=False)
    value_workbook = load_workbook(output_path, data_only=True)
    assert formula_workbook["Results_001"]["B2"].value == "=UPPER(A2)"
    assert value_workbook["Results_001"]["B2"].value == "JOAO SILVA"


def test_csv_output_supports_transforms_and_lookups(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    lookup_path = tmp_path / "ids.csv"
    output_path = tmp_path / "output.csv"
    csv_path.write_text(
        "\n".join(
            [
                "ID,Name,Status",
                "1,Alice,ACTIVE",
                "2,Bob,ACTIVE",
                "2,Bob,ACTIVE",
                "3,Carol,INACTIVE",
            ]
        ),
        encoding="utf-8",
    )
    lookup_path.write_text("ID\n1\n2\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            'ID IN @ids AND Status = "ACTIVE"',
            "--lookup",
            f"ids={lookup_path}:ID",
            "--select",
            "Name",
            "--rename",
            "Name=Person",
            "--dedupe",
            "--sort",
            "Name",
        ],
    )

    assert result.exit_code == 0, result.output
    assert rows_from_csv(output_path) == [("Person",), ("Alice",), ("Bob",)]


def test_filter_supports_dates_regex_and_sheet_splitting(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            'DATA_ADMISSAO BETWEEN "2020-01-01" AND "2023-12-31" AND CIDADE regex "^SAO "',
            "--select",
            "NOME",
            "--max-rows-per-sheet",
            "2",
        ],
    )

    assert result.exit_code == 0, result.output
    workbook = load_workbook(output_path, read_only=True)
    assert "Results_001" in workbook.sheetnames
    assert "Results_002" in workbook.sheetnames


def test_validate_filter_command(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["validate-filter", str(csv_path), "--where", 'STATUS IN ("ATIVO", "SUSPENSO")'])

    assert result.exit_code == 0, result.output
    assert "Filtro válido" in result.output


def test_validate_filter_command_supports_en_us_language(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        ["--idioma", "en-US", "validar-filtro", str(csv_path), "--filtro", 'STATUS IN ("ATIVO", "SUSPENSO")'],
    )

    assert result.exit_code == 0, result.output
    assert "Filter is valid" in result.output


def test_validate_filter_command_accepts_pt_br_command_and_option(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["validar-filtro", str(csv_path), "--filtro", 'STATUS EM ("ATIVO", "SUSPENSO")'])

    assert result.exit_code == 0, result.output
    assert "Filtro válido" in result.output


def test_inspect_command_accepts_pt_br_command_and_option(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["inspecionar", str(csv_path), "--delimitador", ","])

    assert result.exit_code == 0, result.output
    assert "Coluna" in result.output
    assert "CD_EMPRESA" in result.output


def test_cli_help_defaults_to_pt_br() -> None:
    result = runner.invoke(app, ["--help"])

    assert result.exit_code == 0, result.output
    assert "Filtra arquivos grandes" in result.output
    assert "filtrar" in result.output
    assert "validar-filtro" in result.output
    assert "--idioma" in result.output


def test_filter_help_defaults_to_pt_br_option_names() -> None:
    result = runner.invoke(app, ["filtrar", "--help"])

    assert result.exit_code == 0, result.output
    assert "--saida" in result.output
    assert "--filtro" in result.output
    assert "--selecionar" in result.output
    assert "Sem formato" in result.output
    assert "CSV" in result.output
    assert "formato do arquivo de entrada" not in result.output


def test_cli_help_can_be_created_in_en_us() -> None:
    result = runner.invoke(create_app("en-US"), ["--help"])

    assert result.exit_code == 0, result.output
    assert "Filter large files" in result.output


def test_create_app_rejects_invalid_pre_scanned_language() -> None:
    with pytest.raises(ValueError, match="Idioma inválido"):
        create_app("fr-FR")


def test_invalid_cli_language_fails_clearly(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["--idioma", "fr-FR", "validar-filtro", str(csv_path), "--filtro", "CD_EMPRESA = 1"])

    assert result.exit_code == 2
    assert "Idioma inválido" in result.output


def test_validate_filter_rejects_regex_duckdb_cannot_execute(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["validate-filter", str(csv_path), "--where", r'NOME regex "(.)\\1"'])

    assert result.exit_code == 2
    assert "Padrão regex inválido para DuckDB" in result.output


def test_validate_filter_supports_explicit_date_literal(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["validate-filter", str(csv_path), "--where", 'DATA_ADMISSAO = date("2021-05-01")'])

    assert result.exit_code == 0, result.output
    assert "Filtro válido" in result.output


def test_validate_filter_invalid_explicit_date_is_user_facing(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(app, ["validate-filter", str(csv_path), "--where", 'DATA_ADMISSAO = date("2021-99-99")'])

    assert result.exit_code == 2
    assert "Literal de data inválido" in result.output
    assert "Traceback" not in result.output


def test_validate_filter_supports_lookup_from_config(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    lookup_path = tmp_path / "ids.csv"
    config_path = tmp_path / "filters.yaml"
    write_csv(csv_path)
    lookup_path.write_text("CD_EMPRESA\n1\n", encoding="utf-8")
    config_path.write_text(
        f"""
presets:
  empresas:
    where: CD_EMPRESA IN @active_ids
    lookup:
      - active_ids={lookup_path}:CD_EMPRESA
""".strip(),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "validate-filter",
            str(csv_path),
            "--config",
            str(config_path),
            "--preset",
            "empresas",
        ],
    )

    assert result.exit_code == 0, result.output
    assert "Filtro válido" in result.output


def test_filter_resolves_relative_lookup_paths_from_config_directory(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    lookup_path = tmp_path / "ids.csv"
    config_path = tmp_path / "filters.yaml"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)
    lookup_path.write_text("CD_EMPRESA\n2\n", encoding="utf-8")
    config_path.write_text(
        """
presets:
  empresas:
    where: CD_EMPRESA IN @active_ids
    lookup:
      - active_ids=ids.csv:CD_EMPRESA
    select:
      - NOME
""".strip(),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--config",
            str(config_path),
            "--preset",
            "empresas",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("NOME",), ("ANA LIMA",)]


def test_filter_combines_config_and_cli_lookups(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    config_lookup_path = tmp_path / "config_ids.csv"
    cli_lookup_path = tmp_path / "cli_statuses.csv"
    config_path = tmp_path / "filters.yaml"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)
    config_lookup_path.write_text("CD_EMPRESA\n1\n", encoding="utf-8")
    cli_lookup_path.write_text("STATUS\nATIVO\n", encoding="utf-8")
    config_path.write_text(
        """
presets:
  empresas:
    where: CD_EMPRESA IN @config_ids
    lookup:
      - config_ids=config_ids.csv:CD_EMPRESA
    select:
      - NOME
""".strip(),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--config",
            str(config_path),
            "--preset",
            "empresas",
            "--where",
            "STATUS IN @cli_statuses",
            "--lookup",
            f"cli_statuses={cli_lookup_path}:STATUS",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("NOME",), ("JOAO SILVA",), ("MARIA SILVA",)]


def test_filter_supports_config_presets_and_renames(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    config_path = tmp_path / "filters.yaml"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)
    config_path.write_text(
        """
presets:
  ativos:
    where: STATUS = "ATIVO"
    select:
      - NOME
      - STATUS
    rename:
      NOME: Nome
""".strip(),
        encoding="utf-8",
    )

    result = runner.invoke(
        app,
        ["filter", str(csv_path), "-o", str(output_path), "--config", str(config_path), "--preset", "ativos"],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("Nome", "STATUS"), ("JOAO SILVA", "ATIVO"), ("MARIA SILVA", "ATIVO")]


def test_filter_supports_external_lookup(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    lookup_path = tmp_path / "ids.csv"
    output_path = tmp_path / "output.xlsx"
    write_csv(csv_path)
    lookup_path.write_text("CD_EMPRESA\n2\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--where",
            "CD_EMPRESA IN @active_ids",
            "--lookup",
            f"active_ids={lookup_path}:CD_EMPRESA",
            "--select",
            "NOME",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("NOME",), ("ANA LIMA",)]


def test_filter_resolves_lookup_columns_case_insensitively(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    lookup_path = tmp_path / "ids.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("A,Name\n1,Alice\n2,Bob\n", encoding="utf-8")
    lookup_path.write_text("Id\n2\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--case-insensitive-columns",
            "--where",
            "a IN @ids",
            "--lookup",
            f"ids={lookup_path}:id",
            "--select",
            "name",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("Name",), ("Bob",)]


def test_filter_supports_null_literal(tmp_path: Path) -> None:
    csv_path = tmp_path / "nulls.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("A,B\n,empty\n1,one\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--null-value",
            "",
            "--where",
            "A = null",
            "--select",
            "B",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("B",), ("empty",)]


def test_filter_date_like_string_equality_does_not_cast_text_column(tmp_path: Path) -> None:
    csv_path = tmp_path / "codes.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("CODE\nABC\n2024-01-31\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--where",
            'CODE != "2024-01-31"',
            "--select",
            "CODE",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("CODE",), ("ABC",)]


def test_filter_reversed_date_like_ordering_casts_text_column(tmp_path: Path) -> None:
    csv_path = tmp_path / "dates.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("DATA\n2023-12-31\nbad\n2019-01-01\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--where",
            '"2020-01-01" < DATA',
            "--select",
            "DATA",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("DATA",), ("2023-12-31",)]


def test_dedupe_sort_can_use_non_exported_column(tmp_path: Path) -> None:
    csv_path = tmp_path / "dedupe.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("A,B,C\nx,2,k1\ny,1,k2\nx,3,k1\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--where",
            'A != "z"',
            "--select",
            "A",
            "--sort",
            "B",
            "--dedupe-key",
            "C",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("A",), ("y",), ("x",)]


def test_case_insensitive_rename_uses_resolved_column_name(tmp_path: Path) -> None:
    csv_path = tmp_path / "case.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("Name\nAlice\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--case-insensitive-columns",
            "--select",
            "name",
            "--rename",
            "name=FullName",
        ],
    )

    assert result.exit_code == 0, result.output
    rows = rows_from_xlsx(output_path)
    assert rows == [("FullName",), ("Alice",)]


def test_strict_values_cast_errors_are_user_facing(tmp_path: Path) -> None:
    csv_path = tmp_path / "dirty.csv"
    output_path = tmp_path / "output.xlsx"
    csv_path.write_text("A\n1\nnope\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--strict-values",
            "--where",
            "A > 0",
        ],
    )

    assert result.exit_code == 1
    assert "Consulta DuckDB falhou" in result.output
    assert "Traceback" not in result.output


def test_rejects_requires_store_rejects_before_export(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.xlsx"
    rejects_path = tmp_path / "rejects.csv"
    csv_path.write_text("A\n1\n", encoding="utf-8")

    result = runner.invoke(
        app,
        [
            "filter",
            str(csv_path),
            "-o",
            str(output_path),
            "--delimiter",
            ",",
            "--where",
            "A IS NOT NULL",
            "--rejects",
            str(rejects_path),
        ],
    )

    assert result.exit_code == 2
    assert "--rejects requer --store-rejects" in result.output
    assert not output_path.exists()
    assert not rejects_path.exists()


def test_validate_filter_rejects_missing_type_override_column(tmp_path: Path) -> None:
    csv_path = tmp_path / "input.csv"
    write_csv(csv_path)

    result = runner.invoke(
        app,
        [
            "validate-filter",
            str(csv_path),
            "--type",
            "MISSING=decimal",
            "--where",
            "CD_EMPRESA = 1",
        ],
    )

    assert result.exit_code == 2
    assert "Coluna ausente 'MISSING'" in result.output
