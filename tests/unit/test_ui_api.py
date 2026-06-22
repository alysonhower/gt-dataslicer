from pathlib import Path
import json
import time
import zipfile

import duckdb
from openpyxl import load_workbook

from gt_dataslicer.i18n import messages_for
from gt_dataslicer.ui.api import DataSlicerApi
from gt_dataslicer.ui import api as ui_api


def write_people_csv(path: Path) -> None:
    path.write_text(
        "\n".join(
            [
                "ID,Nome,Status,Valor",
                "1,Ana Silva,ATIVO,1500",
                "2,Bruno Lima,CANCELADO,500",
                "3,Camila Silva,ATIVO,2500",
            ]
        ),
        encoding="utf-8",
    )


def wait_for_job(api: DataSlicerApi, job_id: str) -> dict[str, object]:
    for _ in range(100):
        response = api.get_job_status(job_id)
        assert response["ok"], response
        data = response["data"]
        assert isinstance(data, dict)
        if not data["running"]:
            return data
        time.sleep(0.02)
    raise AssertionError("job did not finish")


def test_ui_api_inspects_and_validates_visual_filter(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    inspect_response = api.inspect_csv({"input_path": str(csv_path), "csv_options": {"delimiter": ","}})
    assert inspect_response["ok"], inspect_response
    inspect_data = inspect_response["data"]
    assert isinstance(inspect_data, dict)
    assert inspect_data["columns"] == ["ID", "Nome", "Status", "Valor"]

    validate_response = api.validate_filter(
        {
            "input_path": str(csv_path),
            "csv_options": {"delimiter": ","},
            "filters": {
                "mode": "visual",
                "conditions": [
                    {"column": "Status", "operator": "equals", "value": "ATIVO", "value_type": "string"},
                    {"column": "Valor", "operator": "gt", "value": "1000", "value_type": "number"},
                ],
            },
        }
    )

    assert validate_response["ok"], validate_response
    validate_data = validate_response["data"]
    assert isinstance(validate_data, dict)
    assert validate_data["valid"] is True
    assert "Status" in str(validate_data["sql"])


def test_ui_api_inspect_returns_input_resolution_warnings(tmp_path: Path) -> None:
    zip_path = tmp_path / "inputs.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.writestr("data.csv", "ID,Status\n1,ATIVO\n")
        archive.writestr("ignored.txt", "not supported")
    api = DataSlicerApi()

    response = api.inspect_csv({"input_path": str(zip_path), "csv_options": {"delimiter": ","}})

    assert response["ok"], response
    data = response["data"]
    assert isinstance(data, dict)
    assert data["warnings"] == ["Unsupported ZIP entry skipped: ignored.txt"]


def test_ui_api_runs_csv_export(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    output_path = tmp_path / "filtered.csv"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    start_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(output_path),
            "output_format": "csv",
            "csv_options": {"delimiter": ","},
            "filters": {
                "mode": "visual",
                "conditions": [{"column": "Nome", "operator": "contains", "value": "Silva"}],
            },
            "select": ["Nome"],
        }
    )

    assert start_response["ok"], start_response
    start_data = start_response["data"]
    assert isinstance(start_data, dict)
    job = wait_for_job(api, str(start_data["job_id"]))
    report = job["report"]
    assert isinstance(report, dict)
    assert report["output_rows"] == 2
    assert output_path.read_text(encoding="utf-8").splitlines() == ["Nome", "Ana Silva", "Camila Silva"]


def test_ui_api_reports_single_file_duckdb_progress_without_fake_percent(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    output_path = tmp_path / "filtered.csv"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    start_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(output_path),
            "output_format": "csv",
            "zip_passwords": ["secret-pass"],
            "csv_options": {"delimiter": ","},
            "filters": {
                "mode": "visual",
                "conditions": [{"column": "Nome", "operator": "contains", "value": "Ana Silva"}],
            },
        }
    )

    assert start_response["ok"], start_response
    start_data = start_response["data"]
    assert isinstance(start_data, dict)
    job = wait_for_job(api, str(start_data["job_id"]))
    progress = job["progress"]
    assert isinstance(progress, dict)
    assert job["phase"] == "done"
    assert progress["phase"] == "done"
    assert progress["percent"] == 100
    assert progress["determinate"] is True
    timeline = progress["timeline"]
    assert isinstance(timeline, list)
    engine_entries = [entry for entry in timeline if entry["phase"] in {"inspecting", "validating", "exporting"}]
    assert engine_entries
    assert all(entry["determinate"] is False and entry["percent"] is None for entry in engine_entries)
    progress_json = json.dumps(progress, ensure_ascii=False)
    assert "secret-pass" not in progress_json
    assert "Ana Silva" not in progress_json


def test_ui_api_reports_queue_progress_from_known_item_counts(tmp_path: Path) -> None:
    first_path = tmp_path / "first.csv"
    second_path = tmp_path / "second.csv"
    output_dir = tmp_path / "outputs"
    output_dir.mkdir()
    write_people_csv(first_path)
    write_people_csv(second_path)
    api = DataSlicerApi()

    start_response = api.start_filter_run(
        {
            "input_paths": [str(first_path), str(second_path)],
            "output_path": str(output_dir),
            "output_names": ["first", "second"],
            "avoid_existing_output_paths": True,
            "csv_options": {"delimiter": ","},
            "filters": {"mode": "visual", "conditions": []},
        }
    )

    assert start_response["ok"], start_response
    start_data = start_response["data"]
    assert isinstance(start_data, dict)
    job = wait_for_job(api, str(start_data["job_id"]))
    progress = job["progress"]
    assert isinstance(progress, dict)
    timeline = progress["timeline"]
    assert isinstance(timeline, list)
    queue_entries = [entry for entry in timeline if entry["input_total"] == 2]
    assert queue_entries
    assert any(entry["input_index"] == 1 and entry["percent"] == 0 for entry in queue_entries)
    assert any(entry["input_index"] == 2 and entry["percent"] == 50 for entry in queue_entries)
    assert all(entry["determinate"] is True for entry in queue_entries)
    assert progress["percent"] == 100


def test_ui_api_directory_output_names_avoid_existing_files(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    output_dir = tmp_path / "outputs"
    output_dir.mkdir()
    existing_output = output_dir / "people_tratada.csv"
    existing_output.write_text("old", encoding="utf-8")
    write_people_csv(csv_path)
    api = DataSlicerApi()

    start_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(output_dir),
            "output_names": ["people_tratada"],
            "avoid_existing_output_paths": True,
            "csv_options": {"delimiter": ","},
            "filters": {
                "mode": "visual",
                "conditions": [{"column": "Nome", "operator": "contains", "value": "Silva"}],
            },
            "select": ["Nome"],
        }
    )

    assert start_response["ok"], start_response
    start_data = start_response["data"]
    assert isinstance(start_data, dict)
    job = wait_for_job(api, str(start_data["job_id"]))
    report = job["report"]
    assert isinstance(report, dict)
    safe_output = output_dir / "people_tratada_002.csv"
    assert report["output_paths"] == [str(safe_output)]
    assert existing_output.read_text(encoding="utf-8") == "old"
    assert safe_output.read_text(encoding="utf-8").splitlines() == ["Nome", "Ana Silva", "Camila Silva"]


def test_ui_api_runs_xlsx_export(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    output_path = tmp_path / "filtered.xlsx"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    start_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(output_path),
            "output_format": "xlsx",
            "csv_options": {"delimiter": ","},
            "filters": {
                "mode": "visual",
                "conditions": [{"column": "Status", "operator": "equals", "value": "ATIVO"}],
            },
            "select": ["Nome"],
        }
    )

    assert start_response["ok"], start_response
    start_data = start_response["data"]
    assert isinstance(start_data, dict)
    job = wait_for_job(api, str(start_data["job_id"]))
    assert job["phase"] == "done"
    workbook = load_workbook(output_path, read_only=True)
    rows = [tuple(row) for row in workbook["Results_001"].iter_rows(values_only=True)]
    assert rows == [("Nome",), ("Ana Silva",), ("Camila Silva",)]


def test_ui_api_runs_summarization_and_summarization_only_outputs(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    output_path = tmp_path / "filtered"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    summary_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(tmp_path),
            "output_names": ["filtered"],
            "avoid_existing_output_paths": True,
            "csv_options": {"delimiter": ","},
            "summarization": True,
            "summarization_group_by": ["Status"],
            "summarization_totals": ["Valor"],
            "summarization_output_suffix": "_resumo",
            "summarization_output_format": "csv",
            "filters": {
                "mode": "visual",
                "conditions": [{"column": "Nome", "operator": "contains", "value": "Silva"}],
            },
        }
    )

    assert summary_response["ok"], summary_response
    summary_data = summary_response["data"]
    assert isinstance(summary_data, dict)
    summary_job = wait_for_job(api, str(summary_data["job_id"]))
    summary_report = summary_job["report"]
    assert isinstance(summary_report, dict)
    assert summary_report["output_rows"] == 2
    assert summary_report["output_paths"] == [
        str(output_path.with_suffix(".csv")),
        str(output_path.with_name("filtered_resumo.csv")),
    ]
    assert output_path.with_suffix(".csv").exists()
    assert output_path.with_name("filtered_resumo.csv").exists()

    summary_only_path = tmp_path / "only_summarization"
    summary_only_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(tmp_path),
            "output_names": ["only_summarization"],
            "avoid_existing_output_paths": True,
            "csv_options": {"delimiter": ","},
            "summarization": True,
            "summarization_only": True,
            "summarization_group_by": ["Status"],
            "summarization_totals": ["Valor"],
            "summarization_output_format": "csv",
            "filters": {
                "mode": "visual",
                "combine": "or",
                "conditions": [{"column": "Status", "operator": "equals", "value": "ATIVO"}],
            },
        }
    )
    assert summary_only_response["ok"], summary_only_response
    summary_only_data = summary_only_response["data"]
    assert isinstance(summary_only_data, dict)
    summary_only_job = wait_for_job(api, str(summary_only_data["job_id"]))
    summary_only_report = summary_only_job["report"]
    assert isinstance(summary_only_report, dict)
    assert summary_only_report["output_rows"] == 1
    assert summary_only_report["output_paths"] == [str(summary_only_path.with_suffix(".csv"))]
    assert summary_only_path.with_suffix(".csv").exists()
    assert not summary_only_path.with_name("only_summarization_summarization.csv").exists()
    assert not summary_only_path.with_name("only_summarization_summarization.xlsx").exists()


def test_ui_api_runs_parquet_export_with_derived_column(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    output_path = tmp_path / "filtered.parquet"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    start_response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_path": str(output_path),
            "output_format": "parquet",
            "csv_options": {"delimiter": ","},
            "filters": {
                "mode": "visual",
                "conditions": [{"column": "Status", "operator": "equals", "value": "ATIVO"}],
            },
            "select": ["Nome"],
            "derived_columns": [
                {
                    "source": "Nome",
                    "name": {"suffix": "LIMPO", "separator": "_"},
                    "transforms": [{"operation": "remove_accents"}, {"operation": "uppercase"}],
                }
            ],
        }
    )

    assert start_response["ok"], start_response
    start_data = start_response["data"]
    assert isinstance(start_data, dict)
    job = wait_for_job(api, str(start_data["job_id"]))
    assert job["phase"] == "done"
    rows = duckdb.connect().execute(f"SELECT * FROM read_parquet('{output_path.as_posix()}')").fetchall()
    assert rows == [("Ana Silva", "ANA SILVA"), ("Camila Silva", "CAMILA SILVA")]


def test_ui_api_returns_json_safe_errors(tmp_path: Path) -> None:
    csv_path = tmp_path / "people.csv"
    write_people_csv(csv_path)
    api = DataSlicerApi()

    response = api.start_filter_run(
        {
            "input_path": str(csv_path),
            "output_format": "csv",
            "filters": {"mode": "raw", "raw": "Status = 'ATIVO'"},
        }
    )

    assert response["ok"] is False
    error = response["error"]
    assert isinstance(error, dict)
    assert "Escolha a pasta de destino" in str(error["message"])


def test_ui_api_message_bundle_supports_both_languages() -> None:
    api = DataSlicerApi()

    pt = api.get_messages("pt-BR")
    en = api.get_messages("en-US")

    assert pt["ok"] is True
    assert en["ok"] is True
    assert pt["data"]["messages"]["ui.app_name"] == "DataSlicer"  # type: ignore[index]
    assert en["data"]["messages"]["ui.error.input_required"] == "Choose an input file before continuing."  # type: ignore[index]
    assert pt["data"]["messages"]["ui.error.output_required"] == "Escolha a pasta de destino antes de continuar."  # type: ignore[index]
    assert en["data"]["messages"]["ui.error.output_required"] == "Choose a destination folder before continuing."  # type: ignore[index]
    assert "Use em Agrupar por" in messages_for("pt-BR")["error.summary_total_derived_text"]
    assert "Use it in Group by" in messages_for("en-US")["error.summary_total_derived_text"]


def test_ui_api_saves_reusable_config_without_zip_passwords(tmp_path: Path, monkeypatch) -> None:
    output_path = tmp_path / "config.yaml"

    class FakeFileDialog:
        SAVE = object()

    class FakeWebview:
        FileDialog = FakeFileDialog

    class FakeWindow:
        def create_file_dialog(self, *_args, **_kwargs):
            return (str(output_path),)

    monkeypatch.setattr(ui_api, "_import_webview", lambda: FakeWebview)
    api = DataSlicerApi()
    api.bind_window(FakeWindow())

    response = api.save_config(
        {
            "output_format": "parquet",
            "zip_passwords": ["secret"],
            "filters": {"mode": "raw", "raw": "Status = 'ATIVO'"},
            "select": ["Nome"],
            "derived_columns": [
                {
                    "source": "Nome",
                    "name": {"suffix": "LIMPO", "separator": "_"},
                    "transforms": [{"operation": "uppercase"}],
                }
            ],
        }
    )

    assert response["ok"], response
    contents = output_path.read_text(encoding="utf-8")
    assert "output_format: parquet" in contents
    assert "derived_columns:" in contents
    assert "secret" not in contents


def test_ui_api_saves_summarization_config_with_public_keys(tmp_path: Path, monkeypatch) -> None:
    output_path = tmp_path / "config.yaml"

    class FakeFileDialog:
        SAVE = object()

    class FakeWebview:
        FileDialog = FakeFileDialog

    class FakeWindow:
        def create_file_dialog(self, *_args, **_kwargs):
            return (str(output_path),)

    monkeypatch.setattr(ui_api, "_import_webview", lambda: FakeWebview)
    api = DataSlicerApi()
    api.bind_window(FakeWindow())

    response = api.save_config(
        {
            "summarization": True,
            "summarization_only": True,
            "summarization_group_by": ["Status"],
            "summarization_totals": ["Valor"],
            "summarization_output_suffix": "_resumo",
            "summarization_output_format": "csv",
        }
    )

    assert response["ok"], response
    contents = output_path.read_text(encoding="utf-8")
    assert "summarization: true" in contents
    assert "summarization_only: true" in contents
    assert "summarization_group_by:" in contents
    assert "summarization_totals:" in contents
    assert "summarization_output_suffix: _resumo" in contents
    assert "summarization_output_format: csv" in contents
    assert "summary_group_by" not in contents
    assert "summary_output_suffix" not in contents
    assert "summary_output_format" not in contents
