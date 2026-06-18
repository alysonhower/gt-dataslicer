from pathlib import Path
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from gt_dataslicer.export import excel as excel_module
from gt_dataslicer.export.excel import ExcelExportOptions, export_rows_to_xlsx
from gt_dataslicer.exceptions import ExportLimitError
from gt_dataslicer.report import RunReport


def test_summary_sheet_uses_finalized_report_values(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    report = RunReport(input_path="input.csv")

    def finalize_report() -> None:
        report.rejected_rows = 7

    row_count = export_rows_to_xlsx(
        headers=["A"],
        rows=[["x"]],
        options=ExcelExportOptions(output_path=output_path),
        report=report,
        finalize_report=finalize_report,
    )

    workbook = load_workbook(output_path, read_only=True)
    summary = dict(workbook["_Summary"].iter_rows(values_only=True))
    assert row_count == 1
    assert summary["rejected_rows"] == 7


def test_string_values_are_not_exported_as_formulas_or_urls(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    report = RunReport(input_path="input.csv")

    export_rows_to_xlsx(
        headers=["FormulaLike", "UrlLike"],
        rows=[["=1+1", "https://example.com"]],
        options=ExcelExportOptions(output_path=output_path),
        report=report,
    )

    workbook = load_workbook(output_path, data_only=False, read_only=False)
    worksheet = workbook["Results_001"]
    assert worksheet["A2"].data_type == "s"
    assert worksheet["A2"].value == "=1+1"
    assert worksheet["B2"].data_type == "s"
    assert worksheet["B2"].value == "https://example.com"


def test_decimal_and_large_integer_values_are_exported_as_exact_text(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    report = RunReport(input_path="input.csv")

    export_rows_to_xlsx(
        headers=["Decimal", "LargeInt", "SmallInt"],
        rows=[[Decimal("12345678901234567890.12"), 1234567890123456, 12345]],
        options=ExcelExportOptions(output_path=output_path),
        report=report,
    )

    workbook = load_workbook(output_path, data_only=False, read_only=False)
    worksheet = workbook["Results_001"]
    assert worksheet["A2"].data_type == "s"
    assert worksheet["A2"].value == "12345678901234567890.12"
    assert worksheet["B2"].data_type == "s"
    assert worksheet["B2"].value == "1234567890123456"
    assert worksheet["C2"].data_type == "n"
    assert worksheet["C2"].value == 12345


def test_xlsx_export_rejects_strings_excel_would_truncate(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    report = RunReport(input_path="input.csv")

    with pytest.raises(ExportLimitError, match="Excel supports at most") as exc_info:
        export_rows_to_xlsx(
            headers=["A"],
            rows=[["x" * (excel_module.EXCEL_MAX_STRING_LENGTH + 1)]],
            options=ExcelExportOptions(output_path=output_path),
            report=report,
        )

    assert exc_info.value.code == "excel_string_limit"
    assert exc_info.value.context == {"limit": excel_module.EXCEL_MAX_STRING_LENGTH, "length": 32768}
    assert not output_path.exists()
    assert list(tmp_path.glob(".*.tmp-*")) == []


def test_sheet_prefix_apostrophes_are_sanitized(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    report = RunReport(input_path="input.csv")

    export_rows_to_xlsx(
        headers=["A"],
        rows=[["x"]],
        options=ExcelExportOptions(output_path=output_path, sheet_prefix="'Invalid'"),
        report=report,
    )

    workbook = load_workbook(output_path, read_only=True)
    assert "Invalid_001" in workbook.sheetnames


def test_xlsx_export_rejects_invalid_split_options(tmp_path: Path) -> None:
    report = RunReport(input_path="input.csv")

    with pytest.raises(ExportLimitError, match="--split-mode") as split_info:
        export_rows_to_xlsx(
            headers=["A"],
            rows=[["x"]],
            options=ExcelExportOptions(output_path=tmp_path / "output.xlsx", split_mode="invalid"),
            report=report,
        )

    assert split_info.value.code == "excel_split_mode"
    assert split_info.value.context == {"value": "invalid"}

    with pytest.raises(ExportLimitError, match="--sheets-per-file") as sheets_info:
        export_rows_to_xlsx(
            headers=["A"],
            rows=[["x"]],
            options=ExcelExportOptions(output_path=tmp_path / "output.xlsx", split_mode="both", sheets_per_file=0),
            report=report,
        )

    assert sheets_info.value.code == "excel_sheets_per_file"
    assert sheets_info.value.context == {"min": 1, "value": 0}


def test_xlsx_export_rejects_column_limit_with_structured_error(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    report = RunReport(input_path="input.csv")
    monkeypatch.setattr(excel_module, "EXCEL_MAX_COLUMNS", 1)

    with pytest.raises(ExportLimitError, match="Excel supports at most") as exc_info:
        export_rows_to_xlsx(
            headers=["A", "B"],
            rows=[],
            options=ExcelExportOptions(output_path=tmp_path / "output.xlsx"),
            report=report,
        )

    assert exc_info.value.code == "excel_column_limit"
    assert exc_info.value.context == {"limit": 1, "selected": 2}


def test_xlsx_export_keeps_existing_file_when_workbook_close_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    output_path = tmp_path / "output.xlsx"
    output_path.write_bytes(b"old-content")
    report = RunReport(input_path="input.csv")

    class FakeWorksheet:
        def write_string(self, *_args) -> None:
            pass

        def write(self, *_args) -> None:
            pass

        def write_formula(self, *_args) -> None:
            pass

        def freeze_panes(self, *_args) -> None:
            pass

    class FakeWorkbook:
        def __init__(self, *_args, **_kwargs) -> None:
            pass

        def add_worksheet(self, *_args) -> FakeWorksheet:
            return FakeWorksheet()

        def close(self) -> None:
            raise RuntimeError("close failed")

    monkeypatch.setattr(excel_module.xlsxwriter, "Workbook", FakeWorkbook)

    with pytest.raises(RuntimeError, match="close failed"):
        export_rows_to_xlsx(
            headers=["A"],
            rows=[["x"]],
            options=ExcelExportOptions(output_path=output_path),
            report=report,
        )

    assert output_path.read_bytes() == b"old-content"
    assert list(tmp_path.glob(".*.tmp-*")) == []


def test_split_xlsx_export_commits_final_paths_after_success(tmp_path: Path) -> None:
    output_path = tmp_path / "output.xlsx"
    report = RunReport(input_path="input.csv")

    row_count = export_rows_to_xlsx(
        headers=["A"],
        rows=[["one"], ["two"], ["three"]],
        options=ExcelExportOptions(
            output_path=output_path,
            max_rows_per_sheet=2,
            split_mode="files",
        ),
        report=report,
    )

    expected_paths = [
        tmp_path / "output_001.xlsx",
        tmp_path / "output_002.xlsx",
        tmp_path / "output_003.xlsx",
    ]
    assert row_count == 3
    assert report.output_paths == [str(path) for path in expected_paths]
    assert all(path.exists() for path in expected_paths)
    assert list(tmp_path.glob(".*.tmp-*")) == []
