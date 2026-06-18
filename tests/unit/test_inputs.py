from pathlib import Path
import zipfile

from openpyxl import Workbook
import pyzipper
import pytest
import xlsxwriter

from gt_dataslicer.exceptions import ConfigError, InputReadError, ZipPasswordRequiredError
from gt_dataslicer import inputs as inputs_module
from gt_dataslicer.inputs import (
    InputResolutionOptions,
    InputResolutionSession,
    detect_input_format,
    output_path_for_input,
)


def test_detect_input_format_supports_csv_parquet_and_xlsx() -> None:
    assert detect_input_format(Path("input.csv")) == "csv"
    assert detect_input_format(Path("input.parquet")) == "parquet"
    assert detect_input_format(Path("input.pq")) == "parquet"
    assert detect_input_format(Path("input.xlsx")) == "xlsx"
    assert detect_input_format(Path("input.xls")) is None


def test_input_resolution_requires_at_least_one_file_with_structured_error() -> None:
    with pytest.raises(ConfigError, match="At least one input file") as exc_info:
        with InputResolutionSession([]):
            pass

    assert exc_info.value.code == "input_required"
    assert exc_info.value.context == {}


def test_input_resolution_rejects_missing_file_with_structured_error(tmp_path: Path) -> None:
    missing_path = tmp_path / "missing.csv"

    with pytest.raises(InputReadError, match="Input file not found") as exc_info:
        with InputResolutionSession([missing_path]):
            pass

    assert exc_info.value.code == "input_file_not_found"
    assert exc_info.value.context == {"path": str(missing_path)}


def test_input_resolution_rejects_when_no_supported_files_with_structured_error(tmp_path: Path) -> None:
    text_path = tmp_path / "notes.txt"
    text_path.write_text("not tabular\n", encoding="utf-8")

    with pytest.raises(InputReadError, match="No supported input files") as exc_info:
        with InputResolutionSession([text_path]):
            pass

    assert exc_info.value.code == "no_supported_input"
    assert exc_info.value.context == {}


def test_xlsx_is_staged_from_first_sheet_by_default(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Primeira"
    sheet.append(["A", "B"])
    sheet.append([1, "x"])
    second = workbook.create_sheet("Segunda")
    second.append(["A", "B"])
    second.append([2, "y"])
    xlsx_path = tmp_path / "input.xlsx"
    workbook.save(xlsx_path)

    with InputResolutionSession([xlsx_path]) as session:
        assert len(session.inputs) == 1
        resolved = session.inputs[0]
        assert resolved.format == "csv"
        assert resolved.excel_sheet == "Primeira"
        assert resolved.path.read_text(encoding="utf-8").splitlines() == ["A,B", "1,x"]


def test_xlsx_all_sheets_become_queue_items(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "Aba 1"
    workbook.active.append(["A"])
    workbook.create_sheet("Aba 2").append(["B"])
    xlsx_path = tmp_path / "input.xlsx"
    workbook.save(xlsx_path)

    with InputResolutionSession([xlsx_path], options=InputResolutionOptions(excel_all_sheets=True)) as session:
        assert [item.excel_sheet for item in session.inputs] == ["Aba 1", "Aba 2"]


def test_xlsx_sheet_staging_paths_are_unique_after_sanitizing_sheet_names(tmp_path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "A B"
    first.append(["VAL"])
    first.append(["FIRST"])
    second = workbook.create_sheet("A_B")
    second.append(["VAL"])
    second.append(["SECOND"])
    xlsx_path = tmp_path / "input.xlsx"
    workbook.save(xlsx_path)

    with InputResolutionSession([xlsx_path], options=InputResolutionOptions(excel_all_sheets=True)) as session:
        assert len(session.inputs) == 2
        assert session.inputs[0].path != session.inputs[1].path
        assert session.inputs[0].path.read_text(encoding="utf-8").splitlines() == ["VAL", "FIRST"]
        assert session.inputs[1].path.read_text(encoding="utf-8").splitlines() == ["VAL", "SECOND"]


def test_xlsx_rejects_too_many_selected_sheets(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = Workbook()
    workbook.active.title = "Aba 1"
    workbook.create_sheet("Aba 2")
    xlsx_path = tmp_path / "many-sheets.xlsx"
    workbook.save(xlsx_path)
    monkeypatch.setattr(inputs_module, "XLSX_WORKSHEET_LIMIT", 1)

    with pytest.raises(InputReadError, match="too many worksheets") as exc_info:
        with InputResolutionSession([xlsx_path], options=InputResolutionOptions(excel_all_sheets=True)):
            pass

    assert exc_info.value.code == "excel_too_many_worksheets"
    assert exc_info.value.context == {"count": 2, "limit": 1}


def test_xlsx_rejects_excessive_declared_dimensions(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dados"
    sheet["A1"] = "A"
    sheet["B2"] = "x"
    xlsx_path = tmp_path / "wide-range.xlsx"
    workbook.save(xlsx_path)
    monkeypatch.setattr(inputs_module, "XLSX_STAGED_CELL_LIMIT", 3)

    with pytest.raises(InputReadError, match="declares 2 rows and 2 columns") as exc_info:
        with InputResolutionSession([xlsx_path]):
            pass

    assert exc_info.value.code == "excel_sheet_too_large"
    assert exc_info.value.context["sheet"] == "Dados"
    assert exc_info.value.context["rows"] == 2
    assert exc_info.value.context["columns"] == 2
    assert exc_info.value.context["cells"] == 4
    assert exc_info.value.context["limit"] == 3


def test_xlsx_rejects_large_internal_archive(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    workbook = Workbook()
    workbook.active.append(["A"])
    xlsx_path = tmp_path / "too-large.xlsx"
    workbook.save(xlsx_path)
    monkeypatch.setattr(inputs_module, "XLSX_UNCOMPRESSED_LIMIT_BYTES", 1)

    with pytest.raises(InputReadError, match="too large after decompression") as exc_info:
        with InputResolutionSession([xlsx_path]):
            pass

    assert exc_info.value.code == "excel_too_large"
    assert exc_info.value.context == {"path": str(xlsx_path), "limit": 1}


def test_xlsx_formula_without_cached_value_is_rejected(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dados"
    sheet.append(["VAL", "EMPTY"])
    sheet.append(["=1+1", None])
    xlsx_path = tmp_path / "formula-no-cache.xlsx"
    workbook.save(xlsx_path)

    with pytest.raises(InputReadError, match="formula without a cached value") as exc_info:
        with InputResolutionSession([xlsx_path]):
            pass

    assert exc_info.value.code == "excel_formula_missing_cached_value"
    assert exc_info.value.context["sheet"] == "Dados"
    assert "Dados!A2" in str(exc_info.value.context["examples"])


def test_xlsx_formula_with_cached_value_is_staged_with_warning(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "formula-cache.xlsx"
    workbook = xlsxwriter.Workbook(xlsx_path)
    worksheet = workbook.add_worksheet("Dados")
    worksheet.write_row(0, 0, ["VAL", "BLANK"])
    worksheet.write_formula(1, 0, "=1+1", None, 2)
    workbook.close()

    with InputResolutionSession([xlsx_path]) as session:
        assert session.inputs[0].path.read_text(encoding="utf-8").splitlines() == ["VAL,BLANK", "2,"]
        assert any("cached formula values" in warning for warning in session.warnings)


def test_xlsx_staging_paths_are_unique_for_duplicate_file_names(tmp_path: Path) -> None:
    first_dir = tmp_path / "first"
    second_dir = tmp_path / "second"
    first_dir.mkdir()
    second_dir.mkdir()
    paths: list[Path] = []
    for directory, value in ((first_dir, "FIRST"), (second_dir, "SECOND")):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Dados"
        sheet.append(["VAL"])
        sheet.append([value])
        path = directory / "same.xlsx"
        workbook.save(path)
        paths.append(path)

    with InputResolutionSession(paths) as session:
        assert len(session.inputs) == 2
        assert session.inputs[0].path != session.inputs[1].path
        assert session.inputs[0].path.read_text(encoding="utf-8").splitlines() == ["VAL", "FIRST"]
        assert session.inputs[1].path.read_text(encoding="utf-8").splitlines() == ["VAL", "SECOND"]


def test_zip_extracts_supported_files_and_rejects_unsafe_paths(tmp_path: Path) -> None:
    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.writestr("data.csv", "A\n1\n")
        archive.writestr("ignore.txt", "nope")

    with InputResolutionSession([zip_path]) as session:
        assert len(session.inputs) == 1
        assert session.inputs[0].zip_source == zip_path
        assert session.inputs[0].zip_member == "data.csv"
        assert session.inputs[0].source_label == f"{zip_path}!data.csv"
        assert session.inputs[0].path.read_text(encoding="utf-8") == "A\n1\n"
        assert session.warnings == ["Unsupported ZIP entry skipped: ignore.txt"]

    unsafe_zip = tmp_path / "unsafe.zip"
    with zipfile.ZipFile(unsafe_zip, "w") as archive:
        archive.writestr("../evil.csv", "A\n1\n")

    with pytest.raises(InputReadError, match="Unsafe ZIP entry path") as exc_info:
        with InputResolutionSession([unsafe_zip]):
            pass

    assert exc_info.value.code == "unsafe_zip_entry"
    assert exc_info.value.context == {"member": "../evil.csv"}


def test_zip_delete_after_extract_waits_until_members_resolve_successfully(tmp_path: Path) -> None:
    xlsx_path = tmp_path / "formula-no-cache.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Dados"
    sheet.append(["VAL"])
    sheet.append(["=1+1"])
    workbook.save(xlsx_path)

    zip_path = tmp_path / "input.zip"
    with zipfile.ZipFile(zip_path, "w") as archive:
        archive.write(xlsx_path, "formula-no-cache.xlsx")

    with pytest.raises(InputReadError, match="formula without a cached value") as exc_info:
        with InputResolutionSession(
            [zip_path],
            options=InputResolutionOptions(delete_zip_after_extract=True),
        ):
            pass

    assert exc_info.value.code == "excel_formula_missing_cached_value"
    assert zip_path.exists()


def test_encrypted_zip_requires_session_password(tmp_path: Path) -> None:
    zip_path = tmp_path / "secure.zip"
    with pyzipper.AESZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED, encryption=pyzipper.WZ_AES) as archive:
        archive.setpassword(b"secret")
        archive.writestr("data.csv", "A\n1\n")

    with pytest.raises(ZipPasswordRequiredError) as exc_info:
        with InputResolutionSession([zip_path]):
            pass

    assert exc_info.value.code == "zip_password_required"
    assert exc_info.value.context == {"path": str(zip_path)}

    with InputResolutionSession([zip_path], options=InputResolutionOptions(zip_passwords=["secret"])) as session:
        assert len(session.inputs) == 1


def test_output_path_for_queue_uses_numbered_file_names(tmp_path: Path) -> None:
    first = tmp_path / "first.csv"
    second = tmp_path / "second.csv"
    first.write_text("A\n1\n", encoding="utf-8")
    second.write_text("A\n2\n", encoding="utf-8")

    with InputResolutionSession([first, second]) as session:
        output = output_path_for_input(tmp_path / "filtered.csv", session.inputs[1], index=2, total=2, output_format="csv")

    assert output == tmp_path / "filtered_002_second.csv"


def test_output_path_for_directory_queue_keeps_duplicate_stems_unique(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    output_dir.mkdir()
    first_dir = tmp_path / "first"
    second_dir = tmp_path / "second"
    first_dir.mkdir()
    second_dir.mkdir()
    first = first_dir / "same.csv"
    second = second_dir / "same.csv"
    first.write_text("A\n1\n", encoding="utf-8")
    second.write_text("A\n2\n", encoding="utf-8")

    with InputResolutionSession([first, second]) as session:
        outputs = [
            output_path_for_input(output_dir, input_, index=index, total=2, output_format="csv")
            for index, input_ in enumerate(session.inputs, start=1)
        ]

    assert outputs == [
        output_dir / "001_same_filtered.csv",
        output_dir / "002_same_filtered.csv",
    ]
