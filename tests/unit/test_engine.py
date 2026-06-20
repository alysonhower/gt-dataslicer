from pathlib import Path
import csv

from openpyxl import load_workbook
import pytest

from gt_dataslicer.config import CsvOptions, FilterRunOptions
from gt_dataslicer.derived import parse_derived_columns
from gt_dataslicer.engine.duckdb_engine import DuckDBEngine
from gt_dataslicer.exceptions import FilterValidationError


def rows_from_xlsx(path: Path) -> list[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True)
    worksheet = workbook["Results_001"]
    return [tuple(row) for row in worksheet.iter_rows(values_only=True)]


def test_filter_does_not_count_input_rows_without_report(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    csv_path.write_text("A\n1\n2\n", encoding="utf-8")
    engine = DuckDBEngine()

    def fail_count(*args: object, **kwargs: object) -> int:
        raise AssertionError("_count_rows should not run unless report_path is set")

    monkeypatch.setattr(engine, "_count_rows", fail_count)

    report = engine.run_filter(
        FilterRunOptions(
            input_path=csv_path,
            output_path=output_path,
            csv=CsvOptions(delimiter=","),
            where=["A IS NOT NULL"],
        )
    )

    assert report.input_rows is None
    assert report.output_rows == 2


def test_filter_counts_input_rows_when_report_requested(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    report_path = tmp_path / "report.json"
    csv_path.write_text("A\n1\n2\n", encoding="utf-8")
    engine = DuckDBEngine()
    calls = 0

    def fake_count(*args: object, **kwargs: object) -> int:
        nonlocal calls
        calls += 1
        return 2

    monkeypatch.setattr(engine, "_count_rows", fake_count)

    report = engine.run_filter(
        FilterRunOptions(
            input_path=csv_path,
            output_path=output_path,
            report_path=report_path,
            csv=CsvOptions(delimiter=","),
            where=["A IS NOT NULL"],
        )
    )

    assert calls == 1
    assert report.input_rows == 2


def test_filter_csv_export_finalizes_rejected_rows(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    csv_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.csv"
    csv_path.write_text("A\n1\n2\n", encoding="utf-8")
    engine = DuckDBEngine()
    monkeypatch.setattr(engine, "_rejected_row_count", lambda: 7)

    report = engine.run_filter(
        FilterRunOptions(
            input_path=csv_path,
            output_path=output_path,
            csv=CsvOptions(delimiter=","),
        )
    )

    assert report.rejected_rows == 7


def test_filter_xlsx_summary_sheet_uses_run_report_values(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "output.xlsx"
    input_path.write_text("A\n1\n2\n", encoding="utf-8")

    DuckDBEngine().run_filter(
        FilterRunOptions(
            input_path=input_path,
            output_path=output_path,
            output_format="xlsx",
            where=['A = "1"'],
            report_path=tmp_path / "report.json",
            csv=CsvOptions(delimiter=","),
        )
    )

    workbook = load_workbook(output_path, read_only=True)
    try:
        summary = dict(workbook["_Summary"].iter_rows(values_only=True))
    finally:
        workbook.close()

    assert summary["input_path"] == str(input_path)
    assert summary["input_rows"] == 2
    assert summary["filters"] == 'A = "1"'


def test_filter_exports_summary_alongside_filtered_output(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    filtered_path = tmp_path / "filtered.csv"
    input_path.write_text(
        "\n".join(
            [
                "STATUS,VALOR_TOTAL",
                "ATIVO,10",
                "ATIVO,30",
                "SUSPENSO,20",
            ]
        ),
        encoding="utf-8",
    )
    engine = DuckDBEngine()

    report = engine.run_filter(
        FilterRunOptions(
            input_path=input_path,
            output_path=filtered_path,
            where=['STATUS = "ATIVO"'],
            summarize=True,
            summary_group_by=["STATUS"],
            summary_totals=["VALOR_TOTAL"],
            csv=CsvOptions(delimiter=","),
        )
    )

    summary_path = tmp_path / "filtered_summarization.xlsx"
    assert report.output_paths == [str(filtered_path), str(summary_path)]
    assert report.output_rows == 2
    assert filtered_path.exists()
    assert summary_path.exists()
    with filtered_path.open(newline="", encoding="utf-8") as handle:
        assert list(csv.reader(handle)) == [["STATUS", "VALOR_TOTAL"], ["ATIVO", "10"], ["ATIVO", "30"]]
    assert rows_from_xlsx(summary_path) == [("STATUS", "total_VALOR_TOTAL", "count"), ("ATIVO", 40, 2)]


def test_filter_summary_dedupe_matches_filtered_row_identity(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    filtered_path = tmp_path / "filtered.csv"
    input_path.write_text(
        "\n".join(
            [
                "STATUS,VALOR_TOTAL,OBS",
                "ATIVO,10,primeiro",
                "ATIVO,10,segundo",
            ]
        ),
        encoding="utf-8",
    )

    report = DuckDBEngine().run_filter(
        FilterRunOptions(
            input_path=input_path,
            output_path=filtered_path,
            summarize=True,
            summary_group_by=["STATUS"],
            summary_totals=["VALOR_TOTAL"],
            dedupe=True,
            csv=CsvOptions(delimiter=","),
        )
    )

    summary_path = tmp_path / "filtered_summarization.xlsx"
    assert report.output_rows == 2
    with filtered_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.reader(handle))
    assert rows[0] == ["STATUS", "VALOR_TOTAL", "OBS"]
    assert set(tuple(row) for row in rows[1:]) == {
        ("ATIVO", "10", "primeiro"),
        ("ATIVO", "10", "segundo"),
    }
    assert rows_from_xlsx(summary_path) == [("STATUS", "total_VALOR_TOTAL", "count"), ("ATIVO", 20, 2)]


def test_filter_summary_only_exports_only_summary_output(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "summary_only.xlsx"
    input_path.write_text(
        "\n".join(
            [
                "STATUS,VALOR_TOTAL",
                "ATIVO,10",
                "ATIVO,30",
                "SUSPENSO,20",
            ]
        ),
        encoding="utf-8",
    )
    engine = DuckDBEngine()

    report = engine.run_filter(
        FilterRunOptions(
            input_path=input_path,
            output_path=output_path,
            summarize=True,
            summary_only=True,
            summary_output_path=output_path,
            summary_group_by=["STATUS"],
            summary_totals=["VALOR_TOTAL"],
            csv=CsvOptions(delimiter=","),
        )
    )

    assert report.output_rows == 2
    assert report.output_paths == [str(output_path)]
    assert output_path.exists()
    summary_rows = rows_from_xlsx(output_path)
    assert summary_rows[0] == ("STATUS", "total_VALOR_TOTAL", "count")
    assert set(summary_rows[1:]) == {("ATIVO", 40, 2), ("SUSPENSO", 20, 1)}


def test_filter_summary_groups_by_derived_output_name(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "filtered.csv"
    input_path.write_text(
        "\n".join(
            [
                "CPF,VALOR_TOTAL",
                '"111.111.111-11",10',
                '"11111111111",30',
                '"222.222.222-22",20',
            ]
        ),
        encoding="utf-8",
    )

    report = DuckDBEngine().run_filter(
        FilterRunOptions(
            input_path=input_path,
            output_path=output_path,
            summarize=True,
            summary_group_by=["CPF_DIGITS"],
            summary_totals=["VALOR_TOTAL"],
            derived_columns=parse_derived_columns(
                [
                    {
                        "source": "CPF",
                        "name": {"suffix": "DIGITS", "separator": "_"},
                        "transforms": [{"operation": "keep_digits"}],
                    }
                ]
            ),
            csv=CsvOptions(delimiter=","),
        )
    )

    assert report.output_rows == 3
    summary_rows = rows_from_xlsx(tmp_path / "filtered_summarization.xlsx")
    assert summary_rows[0] == ("CPF_DIGITS", "total_VALOR_TOTAL", "count")
    assert set(summary_rows[1:]) == {("11111111111", 40, 2), ("22222222222", 20, 1)}


def test_filter_summary_rejects_non_numeric_derived_total(tmp_path: Path) -> None:
    input_path = tmp_path / "input.csv"
    output_path = tmp_path / "filtered.csv"
    input_path.write_text("Nome,VALOR_TOTAL\nAna,10\nBruno,20\n", encoding="utf-8")

    with pytest.raises(FilterValidationError, match="coluna derivada 'Nome_UPPER'.*Use em Agrupar por"):
        DuckDBEngine().run_filter(
            FilterRunOptions(
                input_path=input_path,
                output_path=output_path,
                summarize=True,
                summary_totals=["Nome_UPPER"],
                derived_columns=parse_derived_columns(
                    [
                        {
                            "source": "Nome",
                            "name": {"suffix": "UPPER", "separator": "_"},
                            "transforms": [{"operation": "uppercase"}],
                        }
                    ]
                ),
                csv=CsvOptions(delimiter=","),
            )
        )
