"""Input file resolution for CSV, Parquet, XLSX, and ZIP sources."""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from pathlib import Path, PurePosixPath
import re
import shutil
import tempfile
from typing import Callable, Final, Iterable, Literal

from openpyxl import load_workbook
import pyzipper

from .exceptions import ConfigError, InputReadError, ZipPasswordRequiredError
from .filters.compiler import quote_literal


InputFormat = Literal["csv", "parquet", "xlsx"]
OutputFormat = Literal["csv", "parquet", "xlsx"]
SUPPORTED_INPUT_SUFFIXES = {".csv", ".parquet", ".pq", ".xlsx", ".zip"}
ZIP_MEMBER_LIMIT = 10_000
ZIP_UNCOMPRESSED_LIMIT_BYTES = 20 * 1024 * 1024 * 1024
WINDOWS_RESERVED_NAMES: Final = {
    "CON",
    "PRN",
    "AUX",
    "NUL",
    "COM1",
    "COM2",
    "COM3",
    "COM4",
    "COM5",
    "COM6",
    "COM7",
    "COM8",
    "COM9",
    "LPT1",
    "LPT2",
    "LPT3",
    "LPT4",
    "LPT5",
    "LPT6",
    "LPT7",
    "LPT8",
    "LPT9",
}


@dataclass(slots=True)
class ResolvedInput:
    path: Path
    format: InputFormat
    display_name: str
    source_path: Path
    zip_source: Path | None = None
    zip_member: str | None = None
    excel_sheet: str | None = None
    staged: bool = False
    warnings: list[str] = field(default_factory=list)

    @property
    def label(self) -> str:
        if self.excel_sheet:
            return f"{self.display_name} [{self.excel_sheet}]"
        return self.display_name

    @property
    def source_label(self) -> str:
        if self.zip_source is not None and self.zip_member:
            return f"{self.zip_source}!{self.zip_member}"
        return str(self.source_path)


@dataclass(slots=True)
class InputResolutionOptions:
    zip_passwords: list[str] = field(default_factory=list)
    prompt_zip_password: Callable[[Path], str | None] | None = None
    delete_zip_after_extract: bool = False
    excel_all_sheets: bool = False


class InputResolutionSession:
    """Owns temporary files created while resolving inputs."""

    def __init__(
        self,
        input_paths: Iterable[Path],
        *,
        options: InputResolutionOptions | None = None,
    ) -> None:
        self.input_paths = [Path(path) for path in input_paths]
        self.options = options or InputResolutionOptions()
        self._tempdir = tempfile.TemporaryDirectory(prefix="gt-dataslicer-inputs-")
        self.temp_root = Path(self._tempdir.name)
        self.inputs: list[ResolvedInput] = []
        self.warnings: list[str] = []

    def __enter__(self) -> "InputResolutionSession":
        self.inputs = self._resolve_all()
        return self

    def __exit__(self, *_exc: object) -> None:
        self.cleanup()

    def cleanup(self) -> None:
        self._tempdir.cleanup()

    def _resolve_all(self) -> list[ResolvedInput]:
        if not self.input_paths:
            raise ConfigError("At least one input file is required.")

        resolved: list[ResolvedInput] = []
        for path in self.input_paths:
            if not path.exists():
                raise InputReadError(f"Input file not found: {path}")
            if not path.is_file():
                raise InputReadError(f"Input path is not a file: {path}")
            suffix = path.suffix.lower()
            if suffix == ".zip":
                resolved.extend(self._resolve_zip(path))
            else:
                item = self._resolve_regular_file(path, source_path=path)
                if item is not None:
                    resolved.extend(item)

        if not resolved:
            raise InputReadError("No supported input files were found.")
        return resolved

    def _resolve_regular_file(
        self,
        path: Path,
        *,
        source_path: Path,
        zip_source: Path | None = None,
        zip_member: str | None = None,
        display_name: str | None = None,
    ) -> list[ResolvedInput] | None:
        input_format = detect_input_format(path)
        if input_format is None:
            self.warnings.append(f"Unsupported input file skipped: {path.name}")
            return None
        if input_format == "xlsx":
            return self._stage_xlsx(
                path,
                source_path=source_path,
                zip_source=zip_source,
                zip_member=zip_member,
                display_name=display_name or source_path.stem,
            )
        return [
            ResolvedInput(
                path=path,
                format=input_format,
                display_name=display_name or source_path.stem,
                source_path=source_path,
                zip_source=zip_source,
                zip_member=zip_member,
            )
        ]

    def _resolve_zip(self, path: Path) -> list[ResolvedInput]:
        zip_dir = self.temp_root / f"zip_{len(list(self.temp_root.glob('zip_*'))) + 1}_{_safe_name(path.stem)}"
        zip_dir.mkdir(parents=True, exist_ok=True)
        password = self._password_for_zip(path)
        supported_paths = self._extract_supported_members(path, zip_dir, password=password)
        if self.options.delete_zip_after_extract and supported_paths:
            path.unlink()

        resolved: list[ResolvedInput] = []
        for extracted in supported_paths:
            member_name = extracted.relative_to(zip_dir).as_posix()
            item = self._resolve_regular_file(
                extracted,
                source_path=path,
                zip_source=path,
                zip_member=member_name,
                display_name=PurePosixPath(member_name).stem,
            )
            if item is not None:
                resolved.extend(item)
        return resolved

    def _password_for_zip(self, path: Path) -> bytes | None:
        encrypted = _zip_is_encrypted(path)
        if not encrypted:
            return None

        for candidate in self.options.zip_passwords:
            if _zip_password_works(path, candidate):
                return candidate.encode("utf-8")

        if self.options.prompt_zip_password is not None:
            prompted = self.options.prompt_zip_password(path)
            if prompted and _zip_password_works(path, prompted):
                return prompted.encode("utf-8")

        raise ZipPasswordRequiredError(f"ZIP file requires a password: {path}")

    def _extract_supported_members(self, path: Path, target_dir: Path, *, password: bytes | None) -> list[Path]:
        extracted: list[Path] = []
        total_size = 0
        try:
            with pyzipper.AESZipFile(path) as archive:
                infos = archive.infolist()
                if len(infos) > ZIP_MEMBER_LIMIT:
                    raise InputReadError(f"ZIP file has too many entries: {path}")
                for info in infos:
                    if info.is_dir():
                        continue
                    total_size += int(info.file_size)
                    if total_size > ZIP_UNCOMPRESSED_LIMIT_BYTES:
                        raise InputReadError(f"ZIP file is too large after extraction: {path}")
                    member_path = _safe_zip_member_path(target_dir, info.filename)
                    if member_path.suffix.lower() not in SUPPORTED_INPUT_SUFFIXES - {".zip"}:
                        self.warnings.append(f"Unsupported ZIP entry skipped: {info.filename}")
                        continue
                    member_path.parent.mkdir(parents=True, exist_ok=True)
                    with archive.open(info, pwd=password) as source, member_path.open("wb") as target:
                        shutil.copyfileobj(source, target)
                    extracted.append(member_path)
        except ZipPasswordRequiredError:
            raise
        except RuntimeError as exc:
            raise ZipPasswordRequiredError(f"ZIP file requires a password: {path}") from exc
        except InputReadError:
            raise
        except Exception as exc:  # noqa: BLE001
            raise InputReadError(f"Could not extract ZIP file {path}: {exc}") from exc
        return extracted

    def _stage_xlsx(
        self,
        path: Path,
        *,
        source_path: Path,
        zip_source: Path | None,
        zip_member: str | None = None,
        display_name: str | None = None,
    ) -> list[ResolvedInput]:
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise InputReadError(f"Could not read Excel file {path}: {exc}") from exc

        sheet_names = workbook.sheetnames if self.options.excel_all_sheets else workbook.sheetnames[:1]
        input_name = display_name or source_path.stem
        staged_dir = self._next_staging_dir("xlsx", input_name)
        resolved: list[ResolvedInput] = []
        for sheet_name in sheet_names:
            staged_path = staged_dir / f"{_safe_name(sheet_name)}.csv"
            staged_path.parent.mkdir(parents=True, exist_ok=True)
            worksheet = workbook[sheet_name]
            with staged_path.open("w", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                for row in worksheet.iter_rows(values_only=True):
                    values = ["" if value is None else value for value in row]
                    if any(value != "" for value in values):
                        writer.writerow(values)
            resolved.append(
                ResolvedInput(
                    path=staged_path,
                    format="csv",
                    display_name=input_name,
                    source_path=source_path,
                    zip_source=zip_source,
                    zip_member=zip_member,
                    excel_sheet=sheet_name,
                    staged=True,
                )
            )
        workbook.close()
        return resolved

    def _next_staging_dir(self, kind: str, name: str) -> Path:
        root = self.temp_root / kind
        root.mkdir(parents=True, exist_ok=True)
        index = len([path for path in root.iterdir() if path.is_dir()]) + 1
        path = root / f"{index:04d}_{_safe_name(name)}"
        path.mkdir(parents=True, exist_ok=True)
        return path


def detect_input_format(path: Path) -> InputFormat | None:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix in {".parquet", ".pq"}:
        return "parquet"
    if suffix == ".xlsx":
        return "xlsx"
    return None


def default_output_format_for_path(path: Path) -> OutputFormat:
    input_format = detect_input_format(path)
    if input_format is None:
        return "csv"
    return input_format


def default_output_format_for_input(input_: ResolvedInput) -> OutputFormat:
    if input_.zip_member:
        member_format = detect_input_format(Path(PurePosixPath(input_.zip_member).name))
        if member_format is not None:
            return member_format
    source_format = detect_input_format(input_.source_path)
    if source_format is not None:
        return source_format
    return input_.format


def source_expression(input_: ResolvedInput, *, csv_expr: Callable[[Path], str]) -> str:
    if input_.format == "parquet":
        return f"read_parquet({quote_literal(str(input_.path))})"
    return csv_expr(input_.path)


def output_path_for_input(
    base_output: Path,
    input_: ResolvedInput,
    *,
    index: int,
    total: int,
    output_format: str,
    output_name: str | None = None,
    artifact: str = "filtered",
) -> Path:
    suffix = f".{output_format}"
    artifact_suffix = "" if artifact == "filtered" else f"_{_safe_output_stem(artifact, output_format)}"
    if output_name:
        safe_stem = _safe_output_stem(output_name, output_format)
        if artifact != "filtered":
            safe_stem = f"{safe_stem}{artifact_suffix}"
        if base_output.exists() and base_output.is_dir():
            return base_output / f"{safe_stem}{suffix}"
        if not base_output.suffix:
            return base_output / f"{safe_stem}{suffix}"
        return base_output.with_name(f"{safe_stem}{suffix}")

    safe_stem = _safe_name(input_.display_name)
    if input_.excel_sheet:
        safe_stem = f"{safe_stem}_{_safe_name(input_.excel_sheet)}"

    if total == 1:
        if artifact != "filtered":
            if base_output.suffix:
                return base_output.with_name(f"{base_output.stem}{artifact_suffix}{suffix}")
            return base_output.with_name(f"{base_output.name}{artifact_suffix}{suffix}")
        return base_output

    if base_output.exists() and base_output.is_dir():
        return base_output / f"{index:03d}_{safe_stem}{artifact_suffix}{suffix}"
    if not base_output.suffix:
        return base_output / f"{index:03d}_{safe_stem}{artifact_suffix}{suffix}"

    return base_output.with_name(f"{base_output.stem}_{index:03d}_{safe_stem}{artifact_suffix}{suffix}")


def _zip_is_encrypted(path: Path) -> bool:
    with pyzipper.AESZipFile(path) as archive:
        return any(bool(info.flag_bits & 0x1) for info in archive.infolist() if not info.is_dir())


def _zip_password_works(path: Path, password: str) -> bool:
    try:
        with pyzipper.AESZipFile(path) as archive:
            for info in archive.infolist():
                if info.is_dir():
                    continue
                with archive.open(info, pwd=password.encode("utf-8")) as member:
                    member.read(1)
                return True
    except Exception:  # noqa: BLE001
        return False
    return True


def _safe_zip_member_path(root: Path, member_name: str) -> Path:
    member = PurePosixPath(member_name)
    if member.is_absolute() or ".." in member.parts:
        raise InputReadError(f"Unsafe ZIP entry path: {member_name}")
    target = root.joinpath(*member.parts).resolve()
    root_resolved = root.resolve()
    if root_resolved not in target.parents and target != root_resolved:
        raise InputReadError(f"Unsafe ZIP entry path: {member_name}")
    return target


def _safe_name(value: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9._-]+", "_", value).strip("._-")
    if not safe:
        return "input"
    stem = safe.split(".", 1)[0].upper()
    if stem in WINDOWS_RESERVED_NAMES:
        return f"input_{safe}"
    return safe


def _safe_output_stem(value: str, output_format: str) -> str:
    name = value.strip()
    suffix = Path(name).suffix.lower()
    if suffix in {".csv", ".xlsx", ".parquet"} or suffix == f".{output_format}":
        name = name[: -len(suffix)]
    return _safe_name(name)
